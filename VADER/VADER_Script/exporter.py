
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _format_workbook(path: str):
    wb = load_workbook(path)
    ws = wb.active

    # title
    for cell in ws[1]:
        cell.fill      = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # sentiment tag column conditional coloring
    sentiment_col = next(
        (i for i, c in enumerate(ws[1], 1) if c.value == "sentiment_label"), None
    )
    colors = {"Positive": "C6EFCE", "Negative": "FFC7CE", "Neutral": "FFEB9C", "N/A": "F2F2F2"}
    if sentiment_col:
        for row in ws.iter_rows(min_row=2, min_col=sentiment_col, max_col=sentiment_col):
            for cell in row:
                label = str(cell.value) if cell.value else "N/A"
                cell.fill      = PatternFill("solid", start_color=colors.get(label, "FFFFFF"),
                                             end_color=colors.get(label, "FFFFFF"))
                cell.font      = Font(name="Arial", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center")

    # automatic column width
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 60) # (maximum 60)

    ws.freeze_panes = "A2"
    wb.save(path)


def save_results(df: pd.DataFrame, output_path: str):
    """saved as XLSX"""
    df.to_excel(output_path, index=False, engine="openpyxl")
    _format_workbook(output_path)
    print(f"\nsaved to：{output_path}")
